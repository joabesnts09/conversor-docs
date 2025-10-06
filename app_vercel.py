#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aplicação Flask para Vercel - Conversor PDF para XLSM
Versão simplificada para debug
"""

from flask import Flask, request, jsonify, Response
import os
import logging
import io
import threading
import uuid
import time

# Importação condicional para evitar erros no Vercel
try:
    from pdf_to_xlsm_converter import PDFToXLSMConverter
except ImportError as e:
    logging.error(f"Erro ao importar PDFToXLSMConverter: {e}")
    PDFToXLSMConverter = None

# Importações básicas para conversão simples
try:
    import PyPDF2
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    import re
    import tempfile
    import os
except ImportError as e:
    logging.error(f"Erro ao importar dependências básicas: {e}")
    PyPDF2 = None
    openpyxl = None

# Configurar Flask
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sua_chave_secreta_vercel')

# Configurar CORS manualmente para React
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Dicionário para armazenar status das conversões
conversion_status = {}
conversion_data = {}

def process_conversion(task_id, file_data, filename):
    """Processa conversão em background"""
    try:
        logger.info(f"Iniciando process_conversion para task_id: {task_id}")
        
        conversion_status[task_id] = {'progress': 0, 'status': 'processing', 'error': None, 'filename': filename}
        
        conversion_status[task_id]['progress'] = 10
        time.sleep(0.5)
        
        conversion_status[task_id]['progress'] = 30
        
        if PDFToXLSMConverter is not None:
            # Usar conversor completo se disponível
            converter = PDFToXLSMConverter()
            time.sleep(0.5)
            
            conversion_status[task_id]['progress'] = 60
            
            logger.info(f"Criando BytesIO com {len(file_data)} bytes")
            file_buffer = io.BytesIO(file_data)
            file_buffer.name = filename
            logger.info(f"BytesIO criado com sucesso")
            
            xlsm_data = converter.convert_pdf_to_xlsm_in_memory(file_buffer)
        else:
            # Conversão básica usando apenas PyPDF2 e openpyxl
            logger.info("Usando conversão básica (sem pandas)")
            conversion_status[task_id]['progress'] = 60
            
            # Criar arquivo XLSM básico
            wb = Workbook()
            ws = wb.active
            ws.title = "Dados PDF"
            
            # Adicionar cabeçalho
            ws['A1'] = "Data"
            ws['B1'] = "Descrição"
            ws['C1'] = "Valor"
            
            # Estilizar cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Extrair texto do PDF
            file_buffer = io.BytesIO(file_data)
            pdf_reader = PyPDF2.PdfReader(file_buffer)
            
            row = 2
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                lines = text.split('\n')
                
                for line in lines:
                    if line.strip():
                        # Tentar extrair dados básicos
                        parts = line.split()
                        if len(parts) >= 2:
                            ws[f'A{row}'] = parts[0] if parts[0] else ""
                            ws[f'B{row}'] = " ".join(parts[1:-1]) if len(parts) > 2 else parts[1] if len(parts) > 1 else ""
                            ws[f'C{row}'] = parts[-1] if parts else ""
                            row += 1
            
            # Salvar em memória
            output = io.BytesIO()
            wb.save(output)
            xlsm_data = output.getvalue()
            output.close()
        time.sleep(0.5)
        
        conversion_status[task_id]['progress'] = 80
        time.sleep(0.5)
        
        conversion_status[task_id] = {
            'progress': 100, 
            'status': 'completed', 
            'error': None,
            'filename': filename
        }
        conversion_data[task_id] = xlsm_data
        
    except Exception as e:
        logger.error(f"Erro na conversão {task_id}: {str(e)}")
        conversion_status[task_id] = {
            'progress': 0, 
            'status': 'error', 
            'error': str(e)
        }

@app.route('/')
def index():
    """API Status"""
    return jsonify({
        'message': 'Conversor PDF para XLSM API',
        'status': 'running',
        'version': '1.0.0'
    })

@app.route('/health')
def health():
    """Endpoint de saúde"""
    return jsonify({'status': 'ok', 'message': 'Aplicação funcionando'})

@app.route('/test')
def test():
    """Endpoint de teste"""
    return jsonify({
        'message': 'Teste funcionando',
        'flask_version': '2.3.3',
        'python_version': os.sys.version
    })

@app.route('/upload', methods=['POST'])
def upload_file():
    """Upload e conversão de arquivo"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        # Validar tipo de arquivo
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Apenas arquivos PDF são aceitos'}), 400
        
        # Ler dados do arquivo
        file_data = file.read()
        
        # Gerar ID único para a tarefa
        task_id = str(uuid.uuid4())
        
        # Inicializar status
        conversion_status[task_id] = {
            'progress': 0,
            'status': 'queued',
            'filename': file.filename
        }
        
        # Iniciar conversão em background
        thread = threading.Thread(
            target=process_conversion,
            args=(task_id, file_data, file.filename)
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'task_id': task_id,
            'message': 'Conversão iniciada'
        })
        
    except Exception as e:
        logger.error(f"Erro no upload: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/status/<task_id>')
def get_status(task_id):
    """Obter status da conversão"""
    if task_id not in conversion_status:
        return jsonify({'error': 'Tarefa não encontrada'}), 404
    
    status = conversion_status[task_id].copy()
    # Remover dados binários para serialização JSON
    if 'xlsm_data' in status:
        del status['xlsm_data']
    
    return jsonify(status)

@app.route('/download/<task_id>')
def download_file(task_id):
    """Download do arquivo convertido"""
    if task_id not in conversion_status:
        return jsonify({'error': 'Tarefa não encontrada'}), 404
    
    if task_id not in conversion_data:
        return jsonify({'error': 'Arquivo não disponível'}), 404
    
    status = conversion_status[task_id]
    if status['status'] != 'completed':
        return jsonify({'error': 'Conversão não concluída'}), 400
    
    # Preparar nome do arquivo
    original_name = status['filename']
    base_name = os.path.splitext(original_name)[0]
    download_filename = f"{base_name}_convertido.xlsm"
    
    # Retornar arquivo
    return Response(
        conversion_data[task_id],
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{download_filename}"'
        }
    )

# Para Vercel
app = app

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)