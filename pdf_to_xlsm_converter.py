#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor de PDF para XLSM
Converte arquivos PDF com tabelas financeiras para formato Excel (XLSM)
"""

import os
import sys
import pandas as pd
import pdfplumber
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import re
from datetime import datetime
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PDFToXLSMConverter:
    def __init__(self):
        self.setup_logging()
    
    def setup_logging(self):
        """Configura o sistema de logging"""
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        file_handler = logging.FileHandler(log_dir / "converter.log", encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
    
    def extract_tables_from_pdf(self, pdf_path):
        """
        Extrai todas as tabelas de um arquivo PDF
        """
        all_tables = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                logger.info(f"Processando PDF com {len(pdf.pages)} páginas")
                
                for page_num, page in enumerate(pdf.pages, 1):
                    logger.info(f"Processando página {page_num}")
                    
                    # Tentar extrair tabelas da página
                    tables = page.extract_tables()
                    
                    if tables:
                        for table in tables:
                            if table and len(table) > 1:  # Verifica se a tabela tem dados
                                # Filtrar apenas linhas com 3 colunas (Data, Movimentações, Valor)
                                filtered_table = []
                                for row in table:
                                    if row and len(row) >= 3:
                                        # Limpar dados da linha
                                        cleaned_row = [cell.strip() if cell else "" for cell in row[:3]]
                                        # Verificar se a linha tem dados válidos
                                        if any(cleaned_row):
                                            filtered_table.append(cleaned_row)
                                
                                if filtered_table:
                                    all_tables.extend(filtered_table)
                                    logger.info(f"Página {page_num}: {len(filtered_table)} linhas extraídas")
                    
                    # Se não encontrou tabelas, tentar extrair texto e procurar padrões
                    if not tables or not any(tables):
                        logger.info(f"Página {page_num}: Tentando extrair texto para encontrar padrões")
                        text = page.extract_text()
                        if text:
                            lines = text.split('\n')
                            # Detectar tipo de arquivo pelo padrão de data
                            is_mercadopago = self.detect_file_type(text)
                            logger.info(f"Página {page_num}: Detectado formato {'MercadoPago' if is_mercadopago else 'Asaas'}")
                            
                            if is_mercadopago:
                                logger.info(f"Página {page_num}: Extraindo padrões MercadoPago")
                                self.extract_mercadopago_patterns(lines, page_num, all_tables)
                            else:
                                logger.info(f"Página {page_num}: Extraindo padrões Asaas")
                                self.extract_asaas_patterns(lines, page_num, all_tables)
                
                logger.info(f"Total de linhas extraídas: {len(all_tables)}")
                return all_tables
                
        except Exception as e:
            logger.error(f"Erro ao extrair tabelas do PDF: {str(e)}")
            raise
    
    def detect_file_type(self, text):
        """Detecta se é arquivo MercadoPago ou Asaas baseado no padrão de data"""
        # MercadoPago usa DD-MM-YYYY, Asaas usa DD/MM/YYYY
        mercadopago_count = len(re.findall(r'\d{2}-\d{2}-\d{4}', text))
        asaas_count = len(re.findall(r'\d{2}/\d{2}/\d{4}', text))
        
        return mercadopago_count > asaas_count
    
    def extract_mercadopago_patterns(self, lines, page_num, all_tables):
        """Extrai padrões específicos do MercadoPago"""
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
            
            # Pular cabeçalhos
            if line in ['Data Descrição ID da operação Valor Saldo', 'Data', 'Descrição', 'Valor']:
                i += 1
                continue
            
            # Padrão MercadoPago: DD-MM-YYYY descrição ID_operacao R$ valor R$ saldo
            mercadopago_pattern = r'^(\d{2}-\d{2}-\d{4})\s+(.+?)\s+(\d+)\s+(R\$\s*[+-]?\d{1,3}(?:\.\d{3})*(?:,\d{2})?)\s+(R\$\s*[+-]?\d{1,3}(?:\.\d{3})*(?:,\d{2})?)$'
            match = re.match(mercadopago_pattern, line)
            
            if match:
                date, description, operation_id, value, balance = match.groups()
                clean_value = re.sub(r'R\$\s*', '', value).strip()
                
                if self.is_valid_date(date) and self.is_valid_value(clean_value):
                    all_tables.append([date, description, clean_value])
                    logger.info(f"Página {page_num}: MercadoPago extraído: {date} | {description[:30]}... | {clean_value}")
            else:
                # Padrão alternativo: DD-MM-YYYY ID_operacao R$ valor R$ saldo (descrição na linha anterior)
                alt_pattern = r'^(\d{2}-\d{2}-\d{4})\s+(\d+)\s+(R\$\s*[+-]?\d{1,3}(?:\.\d{3})*(?:,\d{2})?)\s+(R\$\s*[+-]?\d{1,3}(?:\.\d{3})*(?:,\d{2})?)$'
                alt_match = re.match(alt_pattern, line)
                
                if alt_match and i > 0:
                    date, operation_id, value, balance = alt_match.groups()
                    description = lines[i-1].strip()  # Descrição na linha anterior
                    clean_value = re.sub(r'R\$\s*', '', value).strip()
                    
                    if self.is_valid_date(date) and self.is_valid_value(clean_value) and description:
                        all_tables.append([date, description, clean_value])
                        logger.info(f"Página {page_num}: MercadoPago (alt) extraído: {date} | {description[:30]}... | {clean_value}")
            
            i += 1
    
    def extract_asaas_patterns(self, lines, page_num, all_tables):
        """Extrai padrões específicos do Asaas"""
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue

            # Pular cabeçalhos
            if line in ['Data Movimentações Valor', 'Data', 'Movimentações', 'Valor']:
                i += 1
                continue

            # Padrão para linha completa: DD/MM/YYYY descrição R$ valor
            date_pattern = r'^(\d{1,2}/\d{1,2}/\d{4})\s+(.+?)\s+(R\$\s*[+-]?\d{1,3}(?:\.\d{3})*(?:,\d{2})?)$'
            match = re.match(date_pattern, line)

            if match:
                date, description, value = match.groups()
                # Limpar o valor (remover R$ e espaços)
                clean_value = re.sub(r'R\$\s*', '', value).strip()
                if clean_value and self.is_valid_date(date) and self.is_valid_value(clean_value):
                    all_tables.append([date, description, clean_value])
                    logger.info(f"Página {page_num}: Linha extraída: {date} | {description[:30]}... | {clean_value}")
            else:
                # Verificar se é uma linha quebrada (descrição na linha atual, valor na próxima)
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    
                    # Padrão: linha atual tem data e descrição, próxima linha tem valor
                    date_desc_pattern = r'^(\d{1,2}/\d{1,2}/\d{4})\s+(.+)$'
                    value_pattern = r'^R\$\s*([+-]?\d{1,3}(?:\.\d{3})*(?:,\d{2})?)$'
                    
                    date_desc_match = re.match(date_desc_pattern, line)
                    value_match = re.match(value_pattern, next_line)
                    
                    if date_desc_match and value_match:
                        date, description = date_desc_match.groups()
                        clean_value = value_match.group(1)
                        
                        if self.is_valid_date(date) and self.is_valid_value(clean_value):
                            all_tables.append([date, description, clean_value])
                            logger.info(f"Página {page_num}: Linha quebrada extraída: {date} | {description[:30]}... | {clean_value}")
                            i += 1  # Pular a próxima linha (valor)
                    else:
                        # Verificar se é uma linha quebrada (descrição sem data, data na próxima linha)
                        if i + 1 < len(lines):
                            next_line = lines[i + 1].strip()
                            
                            # Padrão: linha atual tem descrição, próxima linha tem data e valor
                            desc_pattern = r'^(.+)$'
                            date_value_pattern = r'^(\d{1,2}/\d{1,2}/\d{4})\s+R\$\s*([+-]?\d{1,3}(?:\.\d{3})*(?:,\d{2})?)$'
                            
                            desc_match = re.match(desc_pattern, line)
                            date_value_match = re.match(date_value_pattern, next_line)
                            
                            if desc_match and date_value_match and not re.match(r'^\d{1,2}/\d{1,2}/\d{4}', line):
                                description = desc_match.group(1)
                                date, clean_value = date_value_match.groups()
                                
                                if self.is_valid_date(date) and self.is_valid_value(clean_value):
                                    all_tables.append([date, description, clean_value])
                                    logger.info(f"Página {page_num}: Linha quebrada (desc+data) extraída: {date} | {description[:30]}... | {clean_value}")
                                    i += 1  # Pular a próxima linha (data+valor)
            
            i += 1
    
    def clean_and_validate_data(self, tables):
        """
        Limpa e valida os dados extraídos
        """
        cleaned_data = []
        
        for row in tables:
            if len(row) >= 3:
                data, movimentacoes, valor = row[0], row[1], row[2]
                
                # Validar e limpar data
                if self.is_valid_date(data):
                    # Validar e limpar valor
                    if self.is_valid_value(valor):
                        cleaned_data.append([data, movimentacoes, valor])
        
        logger.info(f"Dados válidos após limpeza: {len(cleaned_data)}")
        return cleaned_data
    
    def is_valid_date(self, date_str):
        """
        Valida se a string é uma data válida
        """
        if not date_str:
            return False
        
        # Padrões de data aceitos
        date_patterns = [
            r'^\d{1,2}/\d{1,2}/\d{4}$',  # DD/MM/YYYY
            r'^\d{1,2}-\d{1,2}-\d{4}$',  # DD-MM-YYYY
            r'^\d{4}-\d{1,2}-\d{1,2}$',  # YYYY-MM-DD
            r'^\d{1,2}\.\d{1,2}\.\d{4}$',  # DD.MM.YYYY
            r'^\d{1,2}/\d{1,2}/\d{2}$',   # DD/MM/YY
            r'^\d{1,2}-\d{1,2}-\d{2}$',   # DD-MM-YY
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, date_str.strip()):
                return True
        
        return False
    
    def is_valid_value(self, value_str):
        """
        Valida se a string é um valor monetário válido
        """
        if not value_str:
            return False
        
        # Remove espaços e caracteres especiais
        cleaned_value = re.sub(r'[^\d,.-]', '', value_str.strip())
        
        # Padrões de valor aceitos
        value_patterns = [
            r'^-?\d{1,3}(\.\d{3})*,\d{2}$',  # R$ 1.000,00
            r'^-?\d+,\d{2}$',                 # 100,00
            r'^-?\d+\.\d{2}$',                # 100.00
            r'^-?\d+$',                       # 100
            r'^-?\d{1,3}(,\d{3})*\.\d{2}$',  # 1,000.00
            r'^-?\d{1,3}(\.\d{3})*$',        # 1.000
        ]
        
        for pattern in value_patterns:
            if re.match(pattern, cleaned_value):
                return True
        
        return False
    
    def convert_to_xlsm_format(self, cleaned_data):
        """
        Converte os dados para o formato XLSM com as colunas especificadas
        """
        # Definir as colunas conforme especificado
        columns = [
            'Data',
            'Cód. Conta Debito',
            'Cód. Conta Credito', 
            'Valor',
            'Cód. Histórico',
            'Complemento Histórico',
            'Inicia Lote',
            'Código Matriz/Filial',
            'Centro de Custo Débito',
            'Centro de Custo Crédito'
        ]
        
        # Criar DataFrame com as colunas vazias
        df_data = []
        
        for row in cleaned_data:
            data, movimentacoes, valor = row
            
            # Criar linha com as colunas especificadas
            new_row = [
                data,                    # Data
                '',                      # Cód. Conta Debito (vazio)
                '',                      # Cód. Conta Credito (vazio)
                valor,                   # Valor
                '',                      # Cód. Histórico (vazio)
                movimentacoes,           # Complemento Histórico
                '',                      # Inicia Lote (vazio)
                '',                      # Código Matriz/Filial (vazio)
                '',                      # Centro de Custo Débito (vazio)
                ''                       # Centro de Custo Crédito (vazio)
            ]
            df_data.append(new_row)
        
        df = pd.DataFrame(df_data, columns=columns)
        logger.info(f"DataFrame criado com {len(df)} linhas e {len(columns)} colunas")
        
        return df
    
    def save_to_xlsm(self, df, output_path):
        """
        Salva o DataFrame como arquivo XLSM com formatação idêntica ao arquivo de referência
        """
        try:
            # Verificar se df é um DataFrame
            if not isinstance(df, pd.DataFrame):
                logger.error(f"Erro: df não é um DataFrame, é {type(df)}")
                raise ValueError(f"Parâmetro df deve ser um DataFrame, recebido: {type(df)}")
            
            logger.info(f"Salvando DataFrame com {len(df)} linhas e {len(df.columns)} colunas")
            # Criar diretório se não existir
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Larguras das colunas do arquivo de referência
            reference_column_widths = {
                'A': 10.7109375,   # Data
                'B': 17.28515625,  # Cód. Conta Debito
                'C': 17.85546875,  # Cód. Conta Credito
                'D': 10.28515625,  # Valor
                'E': 13.42578125,  # Cód. Histórico
                'F': 80.7109375,   # Complemento Histórico
                'G': 17.140625,    # Inicia Lote
                'H': 18.5703125,   # Código Matriz/Filial
                'I': 22.0,         # Centro de Custo Débito
                'J': 22.5703125,   # Centro de Custo Crédito
                'K': 3.28515625,   # (vazio)
                'L': 95.7109375    # Instruções
            }
            
            # Salvar como XLSM
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Criar planilha vazia primeiro
                worksheet = writer.book.create_sheet('Dados')
                
                # Adicionar linhas vazias no início (como no arquivo de referência)
                for row in range(1, 5):
                    for col in range(1, 13):
                        worksheet.cell(row=row, column=col, value="")
                
                # Adicionar cabeçalhos na linha 5
                headers = ['Data', 'Cód. Conta Debito', 'Cód. Conta Credito', 'Valor', 
                          'Cód. Histórico', 'Complemento Histórico', 'Inicia Lote', 
                          'Código Matriz/Filial', 'Centro de Custo Débito', 'Centro de Custo Crédito', '', '']
                
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=5, column=col, value=header)
                
                # Adicionar dados a partir da linha 6
                for row_idx, row_data in enumerate(df.values, 6):
                    for col_idx, value in enumerate(row_data, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Aplicar larguras das colunas do arquivo de referência
                for col_letter, width in reference_column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # Aplicar formatação das células
                from openpyxl.styles import Font, Alignment, Border, Side
                
                # Formatação para cabeçalhos (linha 5)
                header_font = Font(name='Calibri', size=11, bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col in range(1, 11):  # Apenas as 10 primeiras colunas
                    cell = worksheet.cell(row=5, column=col)
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Formatação para dados
                data_font = Font(name='Calibri', size=11)
                data_alignment = Alignment(horizontal='left', vertical='center')
                
                for row in range(6, worksheet.max_row + 1):
                    for col in range(1, 11):  # Apenas as 10 primeiras colunas
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = data_alignment
                
                # Aplicar bordas
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in range(5, worksheet.max_row + 1):
                    for col in range(1, 11):
                        worksheet.cell(row=row, column=col).border = thin_border
                
                # Definir altura da linha dos cabeçalhos
                worksheet.row_dimensions[5].height = 31.5
                
                # Remover a planilha padrão se existir
                if 'Sheet' in writer.book.sheetnames:
                    del writer.book['Sheet']
            
            logger.info(f"Arquivo XLSM salvo em: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao salvar arquivo XLSM: {str(e)}")
            raise
    
    def convert_pdf_to_xlsm(self, pdf_path, output_path):
        """
        Função principal para converter PDF para XLSM
        """
        try:
            logger.info(f"Iniciando conversão: {pdf_path} -> {output_path}")
            
            # 1. Extrair tabelas do PDF
            tables = self.extract_tables_from_pdf(pdf_path)
            
            if not tables:
                raise ValueError("Nenhuma tabela válida encontrada no PDF")
            
            # 2. Limpar e validar dados
            cleaned_data = self.clean_and_validate_data(tables)
            
            if not cleaned_data:
                raise ValueError("Nenhum dado válido encontrado após limpeza")
            
            # 3. Converter para formato XLSM
            df = self.convert_to_xlsm_format(cleaned_data)
            
            # 4. Salvar arquivo XLSM
            self.save_to_xlsm(df, output_path)
            
            logger.info("Conversão concluída com sucesso!")
            return True
            
        except Exception as e:
            logger.error(f"Erro na conversão: {str(e)}")
            raise

    def convert_pdf_to_xlsm_in_memory(self, pdf_file):
        """
        Converte PDF para XLSM em memória e retorna os dados do arquivo
        """
        import io
        import tempfile
        
        try:
            logger.info("Iniciando conversão em memória")
            logger.info(f"Tipo do arquivo: {type(pdf_file)}")
            logger.info(f"Arquivo tem método read: {hasattr(pdf_file, 'read')}")
            
            # Salvar arquivo temporário para processamento
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                # Se for um BytesIO, escrever os dados diretamente
                if hasattr(pdf_file, 'read'):
                    logger.info("Processando como BytesIO")
                    pdf_file.seek(0)  # Voltar ao início
                    file_data = pdf_file.read()  # Ler uma vez só
                    logger.info(f"Dados lidos: {len(file_data)} bytes")
                    temp_file.write(file_data)
                else:
                    logger.info("Processando como Flask FileStorage")
                    # Se for um objeto Flask FileStorage
                    pdf_file.save(temp_file.name)
                temp_path = temp_file.name
                logger.info(f"Arquivo temporário criado: {temp_path}")
            
            try:
                # 1. Extrair tabelas do PDF
                tables = self.extract_tables_from_pdf(temp_path)
                
                if not tables:
                    raise ValueError("Nenhuma tabela válida encontrada no PDF")
                
                # 2. Limpar e validar dados
                cleaned_data = self.clean_and_validate_data(tables)
                
                if not cleaned_data:
                    raise ValueError("Nenhum dado válido encontrado após limpeza")
                
                # 3. Converter para formato XLSM
                df = self.convert_to_xlsm_format(cleaned_data)
                
                # 4. Criar XLSM em memória
                xlsm_data = self.create_xlsm_in_memory(df)
                
                logger.info("Conversão em memória concluída com sucesso!")
                return xlsm_data
                
            finally:
                # Limpar arquivo temporário
                os.unlink(temp_path)
                
        except Exception as e:
            logger.error(f"Erro na conversão em memória: {str(e)}")
            raise

    def create_xlsm_in_memory(self, df):
        """
        Cria arquivo XLSM em memória e retorna os dados
        """
        import io
        
        try:
            # Verificar se df é um DataFrame
            if not isinstance(df, pd.DataFrame):
                logger.error(f"Erro: df não é um DataFrame, é {type(df)}")
                raise ValueError(f"Parâmetro df deve ser um DataFrame, recebido: {type(df)}")
            
            logger.info(f"Criando XLSM em memória com {len(df)} linhas e {len(df.columns)} colunas")
            
            # Criar buffer em memória
            output = io.BytesIO()
            
            # Larguras das colunas do arquivo de referência
            reference_column_widths = {
                'A': 10.7109375,   # Data
                'B': 17.28515625,  # Cód. Conta Debito
                'C': 17.85546875,  # Cód. Conta Credito
                'D': 10.28515625,  # Valor
                'E': 13.42578125,  # Cód. Histórico
                'F': 80.7109375,   # Complemento Histórico
                'G': 17.140625,    # Inicia Lote
                'H': 18.5703125,   # Código Matriz/Filial
                'I': 22.0,         # Centro de Custo Débito
                'J': 22.5703125,   # Centro de Custo Crédito
                'K': 3.28515625,   # (vazio)
                'L': 95.7109375    # Instruções
            }
            
            # Salvar como XLSM em memória
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Criar planilha vazia primeiro
                worksheet = writer.book.create_sheet('Dados')
                
                # Adicionar linhas vazias no início (como no arquivo de referência)
                for row in range(1, 5):
                    for col in range(1, 13):
                        worksheet.cell(row=row, column=col, value="")
                
                # Adicionar cabeçalhos na linha 5
                headers = ['Data', 'Cód. Conta Debito', 'Cód. Conta Credito', 'Valor', 
                          'Cód. Histórico', 'Complemento Histórico', 'Inicia Lote', 
                          'Código Matriz/Filial', 'Centro de Custo Débito', 'Centro de Custo Crédito', '', '']
                
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=5, column=col, value=header)
                
                # Adicionar dados a partir da linha 6
                for row_idx, row_data in enumerate(df.values, 6):
                    for col_idx, value in enumerate(row_data, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Aplicar larguras das colunas do arquivo de referência
                for col_letter, width in reference_column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # Aplicar formatação das células
                from openpyxl.styles import Font, Alignment, Border, Side
                
                # Formatação para cabeçalhos (linha 5)
                header_font = Font(name='Calibri', size=11, bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col in range(1, 11):  # Apenas as 10 primeiras colunas
                    cell = worksheet.cell(row=5, column=col)
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Formatação para dados
                data_font = Font(name='Calibri', size=11)
                data_alignment = Alignment(horizontal='left', vertical='center')
                
                for row in range(6, worksheet.max_row + 1):
                    for col in range(1, 11):  # Apenas as 10 primeiras colunas
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = data_alignment
                
                # Aplicar bordas
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in range(5, worksheet.max_row + 1):
                    for col in range(1, 11):
                        worksheet.cell(row=row, column=col).border = thin_border
                
                # Definir altura da linha dos cabeçalhos
                worksheet.row_dimensions[5].height = 31.5
                
                # Remover a planilha padrão se existir
                if 'Sheet' in writer.book.sheetnames:
                    del writer.book['Sheet']
            
            # Retornar dados do buffer
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Erro ao criar XLSM em memória: {str(e)}")
            raise


class PDFConverterGUI:
    def __init__(self):
        self.converter = PDFToXLSMConverter()
        self.setup_gui()
    
    def setup_gui(self):
        """
        Configura a interface gráfica
        """
        self.root = tk.Tk()
        self.root.title("Conversor PDF para XLSM")
        self.root.geometry("600x400")
        self.root.resizable(True, True)
        
        # Configurar estilo
        style = ttk.Style()
        style.theme_use('clam')
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Título
        title_label = ttk.Label(main_frame, text="Conversor PDF para XLSM", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Seleção de arquivo PDF
        ttk.Label(main_frame, text="Arquivo PDF:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.pdf_path_var = tk.StringVar()
        pdf_entry = ttk.Entry(main_frame, textvariable=self.pdf_path_var, width=50)
        pdf_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=5)
        
        pdf_button = ttk.Button(main_frame, text="Selecionar PDF", 
                               command=self.select_pdf_file)
        pdf_button.grid(row=1, column=2, padx=(5, 0), pady=5)
        
        # Seleção de arquivo de saída
        ttk.Label(main_frame, text="Arquivo de saída:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.output_path_var = tk.StringVar()
        output_entry = ttk.Entry(main_frame, textvariable=self.output_path_var, width=50)
        output_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=5)
        
        output_button = ttk.Button(main_frame, text="Selecionar destino", 
                                  command=self.select_output_file)
        output_button.grid(row=2, column=2, padx=(5, 0), pady=5)
        
        # Botão de conversão
        convert_button = ttk.Button(main_frame, text="Converter PDF para XLSM", 
                                   command=self.convert_file, style='Accent.TButton')
        convert_button.grid(row=3, column=0, columnspan=3, pady=20)
        
        # Barra de progresso
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, 
                                           maximum=100, length=400)
        self.progress_bar.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # Área de log
        ttk.Label(main_frame, text="Log de operações:").grid(row=5, column=0, sticky=tk.W, pady=(20, 5))
        
        # Frame para o log
        log_frame = ttk.Frame(main_frame)
        log_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Text widget para log
        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Configurar peso das linhas
        main_frame.rowconfigure(6, weight=1)
        
        # Informações sobre o conversor
        info_text = """
        Este conversor extrai tabelas de arquivos PDF com 3 colunas (Data, Movimentações, Valor)
        e converte para formato XLSM com as colunas especificadas.
        
        Colunas de saída:
        • Data → Data
        • Movimentações → Complemento Histórico  
        • Valor → Valor
        • Colunas adicionais ficam vazias conforme especificado
        """
        
        info_label = ttk.Label(main_frame, text=info_text, font=('Arial', 9), 
                              foreground='gray', justify=tk.LEFT)
        info_label.grid(row=7, column=0, columnspan=3, pady=(20, 0), sticky=tk.W)
    
    def log_message(self, message):
        """
        Adiciona mensagem ao log da interface
        """
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def select_pdf_file(self):
        """
        Seleciona arquivo PDF
        """
        file_path = filedialog.askopenfilename(
            title="Selecionar arquivo PDF",
            filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")]
        )
        
        if file_path:
            self.pdf_path_var.set(file_path)
            # Sugerir nome do arquivo de saída
            pdf_path = Path(file_path)
            output_path = pdf_path.parent / f"{pdf_path.stem}_convertido.xlsm"
            self.output_path_var.set(str(output_path))
            self.log_message(f"PDF selecionado: {file_path}")
    
    def select_output_file(self):
        """
        Seleciona arquivo de saída
        """
        file_path = filedialog.asksaveasfilename(
            title="Salvar arquivo XLSM",
            defaultextension=".xlsm",
            filetypes=[("Arquivos Excel", "*.xlsm"), ("Todos os arquivos", "*.*")]
        )
        
        if file_path:
            self.output_path_var.set(file_path)
            self.log_message(f"Arquivo de saída selecionado: {file_path}")
    
    def convert_file(self):
        """
        Executa a conversão do arquivo
        """
        pdf_path = self.pdf_path_var.get().strip()
        output_path = self.output_path_var.get().strip()
        
        if not pdf_path:
            messagebox.showerror("Erro", "Por favor, selecione um arquivo PDF")
            return
        
        if not output_path:
            messagebox.showerror("Erro", "Por favor, selecione um arquivo de saída")
            return
        
        if not os.path.exists(pdf_path):
            messagebox.showerror("Erro", "Arquivo PDF não encontrado")
            return
        
        try:
            self.log_message("Iniciando conversão...")
            self.progress_var.set(0)
            
            # Executar conversão
            success = self.converter.convert_pdf_to_xlsm(pdf_path, output_path)
            
            if success:
                self.progress_var.set(100)
                self.log_message("Conversão concluída com sucesso!")
                messagebox.showinfo("Sucesso", f"Arquivo convertido com sucesso!\nSalvo em: {output_path}")
            else:
                self.log_message("Erro na conversão")
                messagebox.showerror("Erro", "Falha na conversão do arquivo")
                
        except Exception as e:
            self.log_message(f"Erro: {str(e)}")
            messagebox.showerror("Erro", f"Erro durante a conversão:\n{str(e)}")
            self.progress_var.set(0)
    
    def run(self):
        """
        Inicia a interface gráfica
        """
        self.root.mainloop()


def main():
    """
    Função principal
    """
    try:
        # Verificar se está sendo executado como script
        if len(sys.argv) > 1:
            # Modo linha de comando
            if len(sys.argv) != 3:
                print("Uso: python pdf_to_xlsm_converter.py <arquivo_pdf> <arquivo_saida.xlsm>")
                sys.exit(1)
            
            pdf_path = sys.argv[1]
            output_path = sys.argv[2]
            
            converter = PDFToXLSMConverter()
            success = converter.convert_pdf_to_xlsm(pdf_path, output_path)
            
            if success:
                print(f"Conversão concluída com sucesso! Arquivo salvo em: {output_path}")
            else:
                print("Erro na conversão")
                sys.exit(1)
        else:
            # Modo interface gráfica
            app = PDFConverterGUI()
            app.run()
            
    except KeyboardInterrupt:
        print("\nOperação cancelada pelo usuário")
        sys.exit(0)
    except Exception as e:
        print(f"Erro: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
